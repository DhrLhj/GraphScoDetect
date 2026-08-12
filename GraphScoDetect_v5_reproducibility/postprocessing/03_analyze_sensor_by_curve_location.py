#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Table: Influence of individual sensor removal on scoliosis identification
across primary curve locations under LOSO evaluation.

Uses EXISTING results only. No retraining.

Primary task:
    Task 2 binary scoliosis identification.

Primary curve-location grouping:
    curve1 -> Thoracic / Lumbar / Thoracolumbar.
    This grouping is independent of curve_number, so double-curve patients are
    included according to their recorded primary curve (curve1).

Evaluation:
    - Subject-level, 3-seed probability ensemble.
    - All sensors: read from completed baseline LOSO predictions.
    - Remove S1/S2/S3/S4: read from completed sensor-ablation LOSO predictions.
    - Within each location group all subjects are scoliosis patients (y_true=1),
      therefore subgroup Accuracy == patient Sensitivity.

Delta definition:
    Delta Acc = Acc(Remove Si) - Acc(All)
    Negative delta -> removal worsens identification.
    Positive delta -> removal improves identification.

Additional audit:
    paired subject bootstrap 95% CI for each delta.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd


GROUPS = ["Thoracic", "Lumbar", "Thoracolumbar"]
GROUP_CN = {
    "Thoracic": "胸弯",
    "Lumbar": "腰弯",
    "Thoracolumbar": "胸腰弯",
}

DISPLAY_CONFIGS = [
    ("All sensors (S1+S2+S3+S4)", "all"),
    ("Remove S1", "remove_s1"),
    ("Remove S2", "remove_s2"),
    ("Remove S3", "remove_s3"),
    ("Remove S4", "remove_s4"),
]


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def parse_prob(x: Any) -> np.ndarray:
    if isinstance(x, str):
        s = x.strip()
        try:
            x = json.loads(s.replace("'", '"'))
        except Exception:
            import ast
            x = ast.literal_eval(s)
    arr = np.asarray(x, dtype=float)
    if arr.ndim != 1:
        raise ValueError(f"Invalid probability: {x!r}")
    return arr


def clean_text(x: Any) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", "", str(x).replace("\u3000", " ")).strip()


def parse_primary_location(x: Any, missing_as_thoracic: bool = False) -> Optional[str]:
    s = clean_text(x)
    if not s or any(k in s for k in ["未注明", "不详", "未知"]):
        return "Thoracic" if missing_as_thoracic else None

    # Thoracolumbar must be checked first because it contains both 胸/腰.
    if "胸腰" in s:
        return "Thoracolumbar"

    # Common Chinese/English variants.
    low = s.lower()
    if "thoracolumbar" in low:
        return "Thoracolumbar"
    if "lumbar" in low and "thoracic" not in low:
        return "Lumbar"
    if "thoracic" in low:
        return "Thoracic"

    if "腰" in s and "胸" not in s:
        return "Lumbar"
    if "胸" in s:
        return "Thoracic"
    return None


def read_subject_rows(paths: Iterable[Path]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for path in paths:
        obj = load_json(path)
        seed_from_path = None
        for part in path.parts:
            if part.startswith("seed_"):
                try:
                    seed_from_path = int(part.replace("seed_", ""))
                except Exception:
                    pass
        for r in obj.get("subject_predictions", []):
            row = dict(r)
            if "prob" not in row and "mean_prob" in row:
                row["prob"] = row["mean_prob"]
            if "prob" not in row:
                raise RuntimeError(f"No subject probability in {path}")
            row["seed"] = int(obj.get("seed", seed_from_path))
            rows.append(row)
    if not rows:
        raise RuntimeError("No subject_predictions found.")
    return rows


def baseline_rows(base_root: Path, task: int, seeds: List[int]) -> List[Dict[str, Any]]:
    paths = []
    for seed in seeds:
        ps = sorted(
            (base_root / "results" / f"task{task}" / "loso" / f"seed_{seed}")
            .glob("fold_*/predictions.json")
        )
        if not ps:
            raise FileNotFoundError(
                f"No baseline LOSO predictions for task={task}, seed={seed}"
            )
        paths.extend(ps)
    return read_subject_rows(paths)


def sensor_rows(
    sensor_root: Path,
    task: int,
    canonical_config: str,
    seeds: List[int],
) -> List[Dict[str, Any]]:
    paths = []
    for seed in seeds:
        ps = sorted(
            (
                sensor_root
                / "results"
                / f"task{task}"
                / "loso"
                / canonical_config
                / f"seed_{seed}"
            ).glob("fold_*/predictions.json")
        )
        if not ps:
            raise FileNotFoundError(
                f"No sensor result: task={task}, config={canonical_config}, seed={seed}"
            )
        paths.extend(ps)
    return read_subject_rows(paths)


def seed_ensemble(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Each subject appears once in the LOSO test set per seed.
    Average subject probabilities across seeds.
    """
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        key = str(r.get("subject_key", "")).strip()
        if not key:
            raise RuntimeError("subject_key missing.")
        grouped[key].append(r)

    out = []
    for key, rs in grouped.items():
        truths = {int(r["y_true"]) for r in rs}
        if len(truths) != 1:
            raise RuntimeError(f"Inconsistent truth labels for {key}: {truths}")

        probs = np.stack([parse_prob(r["prob"]) for r in rs])
        mean_prob = probs.mean(axis=0)
        pred = int(np.argmax(mean_prob))
        first = rs[0]

        out.append({
            "subject_key": key,
            "name": first.get("name", ""),
            "center": first.get("center", ""),
            "curve1": first.get("curve1", ""),
            "curve_number": first.get("curve_number", ""),
            "cobb_angle": first.get("cobb_angle", ""),
            "y_true": int(next(iter(truths))),
            "y_pred": pred,
            "correct": int(pred == int(next(iter(truths)))),
            "n_seeds": len(rs),
            "prob_normal": float(mean_prob[0]),
            "prob_scoliosis": float(mean_prob[1]),
        })
    return pd.DataFrame(out)


def paired_bootstrap_delta(
    all_correct: np.ndarray,
    remove_correct: np.ndarray,
    n_boot: int,
    seed: int,
) -> Tuple[float, float, float]:
    """
    Paired subject bootstrap. Returns delta and percentile 95% CI.
    Delta = remove accuracy - all accuracy.
    """
    all_correct = np.asarray(all_correct, dtype=float)
    remove_correct = np.asarray(remove_correct, dtype=float)
    if len(all_correct) != len(remove_correct):
        raise ValueError("Paired arrays have different lengths")
    n = len(all_correct)
    delta = float(remove_correct.mean() - all_correct.mean())
    if n == 0:
        return np.nan, np.nan, np.nan

    rng = np.random.default_rng(seed)
    idx = rng.integers(0, n, size=(n_boot, n))
    vals = remove_correct[idx].mean(axis=1) - all_correct[idx].mean(axis=1)
    lo, hi = np.percentile(vals, [2.5, 97.5])
    return delta, float(lo), float(hi)


def pct(x: float) -> str:
    return "" if not np.isfinite(x) else f"{100*x:.1f}%"


def pp(x: float) -> str:
    if not np.isfinite(x):
        return ""
    return f"{100*x:+.1f} pp"


def ci_pp(lo: float, hi: float) -> str:
    if not (np.isfinite(lo) and np.isfinite(hi)):
        return ""
    return f"[{100*lo:+.1f}, {100*hi:+.1f}] pp"


def style_xlsx(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color=white)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col in ws.columns:
            letter = col[0].column_letter
            width = max(len(str(c.value or "")) for c in list(col)[:1000]) + 2
            ws.column_dimensions[letter].width = min(max(width, 12), 42)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--base_out_root",
        required=True,
        help="Completed baseline root, e.g. new_label_v5_selected_protocols",
    )
    ap.add_argument(
        "--sensor_out_root",
        required=True,
        help="Completed exact LOSO sensor-ablation root",
    )
    ap.add_argument("--out_dir", required=True)
    ap.add_argument("--seeds", nargs="+", type=int, default=[42, 43, 44])
    ap.add_argument("--bootstrap", type=int, default=10000)
    ap.add_argument("--bootstrap_seed", type=int, default=20260802)
    ap.add_argument(
        "--missing_location_as_thoracic",
        action="store_true",
        help="If explicitly required by the manuscript, assign missing curve1 to thoracic. "
             "Default: exclude missing/ambiguous location.",
    )
    args = ap.parse_args()

    base = Path(args.base_out_root).resolve()
    sensor = Path(args.sensor_out_root).resolve()
    out = Path(args.out_dir).resolve()
    out.mkdir(parents=True, exist_ok=True)

    config_map_path = sensor / "config_maps" / "task2.json"
    if not config_map_path.exists():
        raise FileNotFoundError(
            f"Missing sensor config map: {config_map_path}"
        )

    cmap_obj = load_json(config_map_path)
    cmap = cmap_obj["config_mapping"]

    needed = ["all", "remove_s1", "remove_s2", "remove_s3", "remove_s4"]
    missing = [k for k in needed if k not in cmap]
    if missing:
        raise RuntimeError(f"config_maps/task2.json missing keys: {missing}")

    ensemble: Dict[str, pd.DataFrame] = {}

    # Full baseline is deliberately read from the baseline root rather than
    # retrained sensor results.
    ensemble["all"] = seed_ensemble(baseline_rows(base, 2, args.seeds))

    for key in needed[1:]:
        cc = cmap[key]["canonical_config"]
        ensemble[key] = seed_ensemble(sensor_rows(sensor, 2, cc, args.seeds))

    # Verify exact same subject set across configs for paired comparison.
    subject_sets = {k: set(v["subject_key"].astype(str)) for k, v in ensemble.items()}
    reference = subject_sets["all"]
    mismatch = {
        k: {
            "missing_vs_all": sorted(reference - ss),
            "extra_vs_all": sorted(ss - reference),
        }
        for k, ss in subject_sets.items()
        if ss != reference
    }
    if mismatch:
        raise RuntimeError(
            "Subject sets differ across sensor configurations; paired comparison is invalid.\n"
            + json.dumps(mismatch, ensure_ascii=False, indent=2)
        )

    # Use baseline metadata for group definition, then attach each config's prediction.
    meta = ensemble["all"].copy()
    meta["curve_location"] = meta["curve1"].map(
        lambda x: parse_primary_location(
            x, missing_as_thoracic=args.missing_location_as_thoracic
        )
    )

    # This table is scoliosis identification across patient curve locations.
    patients = meta[meta["y_true"].astype(int) == 1].copy()
    included = patients[patients["curve_location"].isin(GROUPS)].copy()
    excluded = patients[~patients["curve_location"].isin(GROUPS)].copy()

    pred_tables = {
        k: v.set_index("subject_key")[["y_pred", "correct", "prob_scoliosis"]]
        for k, v in ensemble.items()
    }

    # Add all predictions to subject audit.
    audit = included[
        ["subject_key", "name", "center", "curve1", "curve_number", "cobb_angle", "curve_location"]
    ].copy()
    for key in needed:
        tmp = pred_tables[key]
        audit[f"{key}_pred"] = audit["subject_key"].map(tmp["y_pred"])
        audit[f"{key}_correct"] = audit["subject_key"].map(tmp["correct"])
        audit[f"{key}_prob_scoliosis"] = audit["subject_key"].map(tmp["prob_scoliosis"])

    # Three output views:
    # 1) Delta table: All row shows absolute Acc; Remove rows show Delta Acc vs All.
    # 2) Absolute table: every sensor configuration shows its actual subgroup Accuracy.
    # 3) Combined table: removal rows show "absolute Acc (Delta Acc)".
    delta_rows = []
    absolute_rows = []
    combined_rows = []
    numeric_rows = []
    bootstrap_rows = []

    for display, key in DISPLAY_CONFIGS:
        delta_row = {"Sensor configuration": display}
        absolute_row = {"Sensor configuration": display}
        combined_row = {"Sensor configuration": display}
        numrow = {"Sensor configuration": display}

        for gi, group in enumerate(GROUPS):
            members = audit[audit["curve_location"] == group]["subject_key"].tolist()
            n = len(members)
            all_corr = pred_tables["all"].loc[members, "correct"].to_numpy(dtype=float)
            this_corr = pred_tables[key].loc[members, "correct"].to_numpy(dtype=float)

            all_acc = float(all_corr.mean()) if n else np.nan
            acc = float(this_corr.mean()) if n else np.nan

            if key == "all":
                delta = 0.0
                lo = hi = 0.0
                delta_display = pct(acc)
                combined_display = pct(acc)
            else:
                delta, lo, hi = paired_bootstrap_delta(
                    all_corr, this_corr,
                    n_boot=args.bootstrap,
                    seed=args.bootstrap_seed + gi * 100 + needed.index(key),
                )
                delta_display = pp(delta)
                combined_display = f"{pct(acc)} ({pp(delta)})"

            # Original presentation.
            delta_row[group] = delta_display

            # New: actual Accuracy for every configuration.
            absolute_row[group] = pct(acc)

            # New: actual Accuracy + Delta.
            combined_row[group] = combined_display

            # Machine-readable audit values.
            numrow[f"{group}_N"] = n
            numrow[f"{group}_Accuracy"] = acc
            numrow[f"{group}_Accuracy_pct"] = 100.0 * acc if np.isfinite(acc) else np.nan
            numrow[f"{group}_All_Accuracy"] = all_acc
            numrow[f"{group}_All_Accuracy_pct"] = 100.0 * all_acc if np.isfinite(all_acc) else np.nan
            numrow[f"{group}_Delta_Acc"] = delta
            numrow[f"{group}_Delta_Acc_pp"] = 100.0 * delta if np.isfinite(delta) else np.nan
            numrow[f"{group}_Delta_CI95_Low"] = lo
            numrow[f"{group}_Delta_CI95_High"] = hi
            numrow[f"{group}_Delta_CI95_Low_pp"] = 100.0 * lo if np.isfinite(lo) else np.nan
            numrow[f"{group}_Delta_CI95_High_pp"] = 100.0 * hi if np.isfinite(hi) else np.nan

            if key != "all":
                bootstrap_rows.append({
                    "Sensor configuration": display,
                    "Curve location": group,
                    "Curve location CN": GROUP_CN[group],
                    "N": n,
                    "All Accuracy": all_acc,
                    "All Accuracy %": 100 * all_acc,
                    "Removal Accuracy": acc,
                    "Removal Accuracy %": 100 * acc,
                    "Delta Acc (Removal-All)": delta,
                    "Delta Acc pp": 100 * delta,
                    "95% CI low pp": 100 * lo,
                    "95% CI high pp": 100 * hi,
                    "95% CI": ci_pp(lo, hi),
                    "Stable decrease": bool(np.isfinite(hi) and hi < 0),
                    "Stable increase": bool(np.isfinite(lo) and lo > 0),
                })

        delta_rows.append(delta_row)
        absolute_rows.append(absolute_row)
        combined_rows.append(combined_row)
        numeric_rows.append(numrow)

    delta_df = pd.DataFrame(delta_rows)
    absolute_df = pd.DataFrame(absolute_rows)
    combined_df = pd.DataFrame(combined_rows)
    num_df = pd.DataFrame(numeric_rows)
    boot_df = pd.DataFrame(bootstrap_rows)

    rename_cn = {
        "Sensor configuration": "传感器配置",
        "Thoracic": "胸弯",
        "Lumbar": "腰弯",
        "Thoracolumbar": "胸腰弯",
    }
    delta_cn_df = delta_df.rename(columns=rename_cn).copy()
    absolute_cn_df = absolute_df.rename(columns=rename_cn).copy()
    combined_cn_df = combined_df.rename(columns=rename_cn).copy()

    # Group counts.
    group_counts = (
        audit.groupby("curve_location")
        .size()
        .reindex(GROUPS, fill_value=0)
        .rename("N")
        .reset_index()
    )
    group_counts["弯曲位置"] = group_counts["curve_location"].map(GROUP_CN)

    excluded_out = excluded[
        ["subject_key", "name", "center", "curve1", "curve_number", "cobb_angle"]
    ].copy()
    excluded_out["reason"] = "Missing/ambiguous primary curve location"

    # CSV outputs.
    delta_df.to_csv(
        out / "Table_sensor_curve_location_LOSO_delta.csv",
        index=False, encoding="utf-8-sig"
    )
    absolute_df.to_csv(
        out / "Table_sensor_curve_location_LOSO_absolute.csv",
        index=False, encoding="utf-8-sig"
    )
    combined_df.to_csv(
        out / "Table_sensor_curve_location_LOSO_combined.csv",
        index=False, encoding="utf-8-sig"
    )

    delta_cn_df.to_csv(
        out / "Table_sensor_curve_location_LOSO_delta_CN.csv",
        index=False, encoding="utf-8-sig"
    )
    absolute_cn_df.to_csv(
        out / "Table_sensor_curve_location_LOSO_absolute_CN.csv",
        index=False, encoding="utf-8-sig"
    )
    combined_cn_df.to_csv(
        out / "Table_sensor_curve_location_LOSO_combined_CN.csv",
        index=False, encoding="utf-8-sig"
    )

    num_df.to_csv(out / "Numeric_Audit.csv", index=False)
    boot_df.to_csv(out / "Paired_Bootstrap_Delta.csv", index=False)
    audit.to_csv(out / "Subject_Details.csv", index=False, encoding="utf-8-sig")
    excluded_out.to_csv(out / "Excluded_Patients.csv", index=False, encoding="utf-8-sig")
    group_counts.to_csv(out / "Group_Counts.csv", index=False, encoding="utf-8-sig")

    readme = pd.DataFrame([
        ["Task", "Task 2 binary scoliosis identification"],
        ["Protocol", "LOSO"],
        ["Evaluation unit", "Subject-level 3-seed probability ensemble"],
        ["Full model source", str(base)],
        ["Sensor-ablation source", str(sensor)],
        ["Grouping", "Primary curve location from curve1; double-curve patients are included according to curve1"],
        ["Missing curve1", "Assigned to thoracic" if args.missing_location_as_thoracic else "Excluded"],
        ["Delta definition", "Delta Acc = Acc(Remove Si) - Acc(All)"],
        ["Absolute table", "Actual subgroup Accuracy for every sensor configuration"],
        ["Combined table", "Actual subgroup Accuracy followed by Delta Acc in parentheses"],
        ["Interpretation", "Negative delta means sensor removal decreases identification performance"],
        ["Important binary note", "All location groups are scoliosis patients, so subgroup Accuracy equals patient Sensitivity"],
        ["Bootstrap", f"Paired subject bootstrap, {args.bootstrap} resamples, percentile 95% CI"],
    ], columns=["Item", "Definition"])

    xlsx = out / "Table_sensor_curve_location_LOSO.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as w:
        delta_df.to_excel(w, sheet_name="Delta_Table_EN", index=False)
        delta_cn_df.to_excel(w, sheet_name="Delta_Table_CN", index=False)

        absolute_df.to_excel(w, sheet_name="Absolute_Acc_EN", index=False)
        absolute_cn_df.to_excel(w, sheet_name="Absolute_Acc_CN", index=False)

        combined_df.to_excel(w, sheet_name="Acc_and_Delta_EN", index=False)
        combined_cn_df.to_excel(w, sheet_name="Acc_and_Delta_CN", index=False)

        group_counts.to_excel(w, sheet_name="Group_Counts", index=False)
        num_df.to_excel(w, sheet_name="Numeric_Audit", index=False)
        boot_df.to_excel(w, sheet_name="Paired_Bootstrap", index=False)
        audit.to_excel(w, sheet_name="Subject_Details", index=False)
        excluded_out.to_excel(w, sheet_name="Excluded_Patients", index=False)
        readme.to_excel(w, sheet_name="README", index=False)

    style_xlsx(xlsx)

    print("\n===== DELTA TABLE =====")
    print(delta_df.to_string(index=False))

    print("\n===== ABSOLUTE ACCURACY TABLE =====")
    print(absolute_df.to_string(index=False))

    print("\n===== ACCURACY + DELTA TABLE =====")
    print(combined_df.to_string(index=False))

    print("\n===== GROUP COUNTS =====")
    print(group_counts.to_string(index=False))

    print("\nDelta definition: Acc(Remove Si) - Acc(All); negative = performance drop.")
    print(f"\n[DONE] {xlsx}")


if __name__ == "__main__":
    main()
