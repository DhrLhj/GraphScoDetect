#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Task-4 subject-level misclassification analysis across ALL experiment protocols.

Protocols:
    gkf3, gkf5, gkf7, gkf10, loso, loco

Primary input:
    summary/07_subject_ensemble_predictions.csv

Optional seed-level input:
    summary/08_subject_predictions_by_seed.csv

No retraining is required.

Core outputs:
1) one row per misclassification EVENT:
   subject + protocol + true class + predicted class
2) one row per SUBJECT:
   how many protocols were wrong, which protocols, and what each protocol predicted
"""

from __future__ import annotations

import argparse
import ast
import json
from collections import Counter
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd


PROTOCOL_ORDER = ["gkf3", "gkf5", "gkf7", "gkf10", "loso", "loco"]

LABEL_NAMES = {
    0: "Normal",
    1: "Mild",
    2: "Moderate",
    3: "Severe",
}


def parse_prob(x: Any) -> Optional[np.ndarray]:
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass

    if isinstance(x, (list, tuple, np.ndarray)):
        arr = np.asarray(x, dtype=float)
        return arr if arr.ndim == 1 else None

    s = str(x).strip()
    if not s:
        return None

    for fn in [json.loads, ast.literal_eval]:
        try:
            arr = np.asarray(fn(s), dtype=float)
            if arr.ndim == 1:
                return arr
        except Exception:
            pass
    return None


def first_existing(df: pd.DataFrame, cols):
    for c in cols:
        if c in df.columns:
            return c
    return None


def add_derived_columns(df: pd.DataFrame) -> pd.DataFrame:
    x = df.copy()

    x["protocol"] = x["protocol"].astype(str).str.lower()
    x["y_true"] = x["y_true"].astype(int)
    x["y_pred"] = x["y_pred"].astype(int)

    x["true_label"] = x["y_true"].map(LABEL_NAMES)
    x["pred_label"] = x["y_pred"].map(LABEL_NAMES)
    x["correct"] = (x["y_true"] == x["y_pred"]).astype(int)

    x["error_transition"] = np.where(
        x["correct"].eq(1),
        "Correct",
        x["true_label"].astype(str) + " -> " + x["pred_label"].astype(str),
    )

    prob_col = first_existing(x, ["ensemble_prob", "prob", "mean_prob"])

    pred_conf = []
    true_prob = []
    pred_prob = []
    margin = []

    for _, r in x.iterrows():
        p = parse_prob(r[prob_col]) if prob_col else None
        if p is None or len(p) < 4:
            pred_conf.append(np.nan)
            true_prob.append(np.nan)
            pred_prob.append(np.nan)
            margin.append(np.nan)
            continue

        yt = int(r["y_true"])
        yp = int(r["y_pred"])

        pred_conf.append(float(np.max(p)))
        true_prob.append(float(p[yt]))
        pred_prob.append(float(p[yp]))
        margin.append(float(p[yp] - p[yt]))

    x["pred_confidence"] = pred_conf
    x["true_class_probability"] = true_prob
    x["predicted_class_probability"] = pred_prob
    x["error_probability_margin"] = margin

    return x


def subject_column(df: pd.DataFrame) -> str:
    c = first_existing(df, ["subject_key", "subject_id", "name"])
    if c is None:
        raise RuntimeError("Cannot find subject_key / subject_id / name.")
    return c


def ordered_protocol_categorical(series: pd.Series):
    return pd.Categorical(
        series,
        categories=PROTOCOL_ORDER,
        ordered=True,
    )


def all_error_events(task4: pd.DataFrame) -> pd.DataFrame:
    wrong = task4[task4["correct"].eq(0)].copy()

    preferred = [
        "protocol",
        "subject_id",
        "subject_key",
        "name",
        "center",
        "cobb_angle",
        "curve_number",
        "curve1",
        "y_true",
        "true_label",
        "y_pred",
        "pred_label",
        "error_transition",
        "pred_confidence",
        "true_class_probability",
        "predicted_class_probability",
        "error_probability_margin",
        "n_seeds",
        "n_samples",
        "n_segments",
        "ensemble_prob",
    ]
    cols = [c for c in preferred if c in wrong.columns]

    wrong["_protocol_order"] = ordered_protocol_categorical(wrong["protocol"])
    sort_cols = ["_protocol_order", "true_label", "pred_label"]
    if "center" in wrong.columns:
        sort_cols.append("center")
    if "name" in wrong.columns:
        sort_cols.append("name")

    wrong = wrong.sort_values(sort_cols).drop(columns=["_protocol_order"])
    return wrong[cols]


def protocol_summary(task4: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for protocol in PROTOCOL_ORDER:
        g = task4[task4["protocol"].eq(protocol)]
        if g.empty:
            rows.append({
                "protocol": protocol,
                "n_subjects": 0,
                "n_correct": 0,
                "n_wrong": 0,
                "accuracy": np.nan,
                "error_rate": np.nan,
            })
            continue

        rows.append({
            "protocol": protocol,
            "n_subjects": len(g),
            "n_correct": int(g["correct"].sum()),
            "n_wrong": int((1 - g["correct"]).sum()),
            "accuracy": float(g["correct"].mean()),
            "error_rate": float(1 - g["correct"].mean()),
        })
    return pd.DataFrame(rows)


def transition_by_protocol(task4: pd.DataFrame) -> pd.DataFrame:
    wrong = task4[task4["correct"].eq(0)].copy()
    if wrong.empty:
        return pd.DataFrame()

    g = (
        wrong.groupby(
            ["protocol", "true_label", "pred_label", "error_transition"],
            observed=True,
        )
        .size()
        .rename("n_errors")
        .reset_index()
    )

    total = (
        wrong.groupby("protocol")
        .size()
        .rename("protocol_total_errors")
        .reset_index()
    )

    g = g.merge(total, on="protocol", how="left")
    g["fraction_within_protocol_errors"] = (
        g["n_errors"] / g["protocol_total_errors"]
    )

    g["_protocol_order"] = ordered_protocol_categorical(g["protocol"])
    return (
        g.sort_values(
            ["_protocol_order", "n_errors"],
            ascending=[True, False],
        )
        .drop(columns=["_protocol_order"])
    )


def class_error_by_protocol(task4: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for protocol in PROTOCOL_ORDER:
        pg = task4[task4["protocol"].eq(protocol)]
        if pg.empty:
            continue

        for cid, cname in LABEL_NAMES.items():
            g = pg[pg["y_true"].eq(cid)]
            n = len(g)
            if n == 0:
                rows.append({
                    "protocol": protocol,
                    "true_class": cname,
                    "n_subjects": 0,
                    "n_correct": 0,
                    "n_wrong": 0,
                    "accuracy": np.nan,
                    "error_rate": np.nan,
                    "most_common_wrong_class": "",
                    "most_common_wrong_count": 0,
                })
                continue

            wrong = g[g["correct"].eq(0)]
            counts = Counter(wrong["pred_label"].tolist())

            rows.append({
                "protocol": protocol,
                "true_class": cname,
                "n_subjects": n,
                "n_correct": int(g["correct"].sum()),
                "n_wrong": len(wrong),
                "accuracy": float(g["correct"].mean()),
                "error_rate": float(1 - g["correct"].mean()),
                "most_common_wrong_class": (
                    counts.most_common(1)[0][0] if counts else ""
                ),
                "most_common_wrong_count": (
                    counts.most_common(1)[0][1] if counts else 0
                ),
            })

    return pd.DataFrame(rows)


def subject_cross_protocol_summary(task4: pd.DataFrame) -> pd.DataFrame:
    sid_col = subject_column(task4)
    rows = []

    for sid, g in task4.groupby(sid_col, dropna=False):
        first = g.iloc[0]

        wrong = g[g["correct"].eq(0)].copy()
        transitions = Counter(wrong["error_transition"].tolist())
        pred_wrong_classes = Counter(wrong["pred_label"].tolist())

        row = {
            "subject_key": str(sid),
            "subject_id": first.get("subject_id", ""),
            "name": first.get("name", ""),
            "center": first.get("center", ""),
            "cobb_angle": first.get("cobb_angle", np.nan),
            "curve_number": first.get("curve_number", ""),
            "curve1": first.get("curve1", ""),
            "true_label": first.get("true_label", ""),
            "n_protocols_available": len(g),
            "n_protocols_correct": int(g["correct"].sum()),
            "n_protocols_wrong": len(wrong),
            "cross_protocol_error_rate": (
                len(wrong) / len(g) if len(g) else np.nan
            ),
            "wrong_protocols": " | ".join(
                p for p in PROTOCOL_ORDER
                if p in set(wrong["protocol"].tolist())
            ),
            "most_common_error_transition": (
                transitions.most_common(1)[0][0]
                if transitions else ""
            ),
            "most_common_error_count": (
                transitions.most_common(1)[0][1]
                if transitions else 0
            ),
            "stable_same_wrong_class": (
                len(wrong) >= 2 and len(pred_wrong_classes) == 1
            ),
        }

        # One explicit prediction/error column for every protocol.
        for protocol in PROTOCOL_ORDER:
            pg = g[g["protocol"].eq(protocol)]
            if pg.empty:
                row[f"{protocol}_pred"] = ""
                row[f"{protocol}_result"] = "NA"
                row[f"{protocol}_transition"] = ""
                row[f"{protocol}_confidence"] = np.nan
            else:
                r = pg.iloc[0]
                row[f"{protocol}_pred"] = r["pred_label"]
                row[f"{protocol}_result"] = (
                    "Correct" if int(r["correct"]) == 1 else "Wrong"
                )
                row[f"{protocol}_transition"] = (
                    "" if int(r["correct"]) == 1 else r["error_transition"]
                )
                row[f"{protocol}_confidence"] = r["pred_confidence"]

        rows.append(row)

    out = pd.DataFrame(rows)

    return out.sort_values(
        ["n_protocols_wrong", "cross_protocol_error_rate", "true_label"],
        ascending=[False, False, True],
    )


def seed_error_events(
    seed_csv: Path,
    task: int = 4,
) -> pd.DataFrame:
    """
    Optional audit table: one row per subject × protocol × seed that is wrong.
    """
    if not seed_csv.exists():
        return pd.DataFrame()

    df = pd.read_csv(seed_csv)
    if "task" in df.columns:
        df = df[df["task"].astype(int).eq(task)].copy()

    if "protocol" not in df.columns:
        return pd.DataFrame()

    df["protocol"] = df["protocol"].astype(str).str.lower()
    df = df[df["protocol"].isin(PROTOCOL_ORDER)].copy()

    if df.empty:
        return pd.DataFrame()

    df["y_true"] = df["y_true"].astype(int)
    df["y_pred"] = df["y_pred"].astype(int)
    df["true_label"] = df["y_true"].map(LABEL_NAMES)
    df["pred_label"] = df["y_pred"].map(LABEL_NAMES)
    df["correct"] = (df["y_true"] == df["y_pred"]).astype(int)
    df["error_transition"] = (
        df["true_label"].astype(str)
        + " -> "
        + df["pred_label"].astype(str)
    )

    wrong = df[df["correct"].eq(0)].copy()

    preferred = [
        "protocol",
        "seed",
        "subject_id",
        "subject_key",
        "name",
        "center",
        "cobb_angle",
        "y_true",
        "true_label",
        "y_pred",
        "pred_label",
        "error_transition",
    ]
    cols = [c for c in preferred if c in wrong.columns]

    wrong["_protocol_order"] = ordered_protocol_categorical(wrong["protocol"])
    sort_cols = ["_protocol_order"]
    if "seed" in wrong.columns:
        sort_cols.append("seed")

    return (
        wrong.sort_values(sort_cols)
        .drop(columns=["_protocol_order"])[cols]
    )


def center_protocol_summary(task4: pd.DataFrame) -> pd.DataFrame:
    if "center" not in task4.columns:
        return pd.DataFrame()

    rows = []
    for (protocol, center), g in task4.groupby(
        ["protocol", "center"],
        dropna=False,
    ):
        rows.append({
            "protocol": protocol,
            "center": center,
            "n_subjects": len(g),
            "n_correct": int(g["correct"].sum()),
            "n_wrong": int((1 - g["correct"]).sum()),
            "accuracy": float(g["correct"].mean()),
            "error_rate": float(1 - g["correct"].mean()),
        })

    out = pd.DataFrame(rows)
    out["_protocol_order"] = ordered_protocol_categorical(out["protocol"])
    return (
        out.sort_values(
            ["_protocol_order", "error_rate"],
            ascending=[True, False],
        )
        .drop(columns=["_protocol_order"])
    )


def style_xlsx(path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for c in ws[1]:
            c.fill = fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.border = Border(
                    left=thin, right=thin,
                    top=thin, bottom=thin
                )
                c.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        for col in ws.columns:
            letter = col[0].column_letter
            mx = max(len(str(c.value or "")) for c in list(col)[:1000])
            ws.column_dimensions[letter].width = min(max(mx + 2, 12), 45)

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()

    ap.add_argument(
        "--summary_dir",
        required=True,
        help=".../new_label_v5_selected_protocols/summary",
    )
    ap.add_argument(
        "--out_dir",
        required=True,
    )
    ap.add_argument(
        "--task",
        type=int,
        default=4,
        choices=[4],
    )
    ap.add_argument(
        "--protocols",
        nargs="+",
        default=PROTOCOL_ORDER,
    )
    ap.add_argument(
        "--frequent_error_min",
        type=int,
        default=2,
        help="At least this many protocols wrong => frequent-error subject.",
    )

    args = ap.parse_args()

    summary = Path(args.summary_dir).resolve()
    out = Path(args.out_dir).resolve()
    out.mkdir(parents=True, exist_ok=True)

    ensemble_csv = summary / "07_subject_ensemble_predictions.csv"
    seed_csv = summary / "08_subject_predictions_by_seed.csv"

    if not ensemble_csv.exists():
        raise FileNotFoundError(ensemble_csv)

    df = pd.read_csv(ensemble_csv)

    if "task" not in df.columns or "protocol" not in df.columns:
        raise RuntimeError(
            "07_subject_ensemble_predictions.csv must contain task and protocol."
        )

    df = df[df["task"].astype(int).eq(args.task)].copy()
    df["protocol"] = df["protocol"].astype(str).str.lower()

    requested = [str(p).lower() for p in args.protocols]
    df = df[df["protocol"].isin(requested)].copy()

    if df.empty:
        raise RuntimeError(
            f"No Task-4 data for protocols: {requested}"
        )

    df = add_derived_columns(df)

    # Sanity check protocol availability.
    protocol_counts = (
        df.groupby("protocol")
        .size()
        .reindex(requested, fill_value=0)
    )
    print("\n===== PROTOCOL SUBJECT COUNTS =====")
    print(protocol_counts.to_string())

    events = all_error_events(df)
    psummary = protocol_summary(df)
    transitions = transition_by_protocol(df)
    class_summary = class_error_by_protocol(df)
    center_summary = center_protocol_summary(df)
    subject_summary = subject_cross_protocol_summary(df)

    frequent = subject_summary[
        subject_summary["n_protocols_wrong"].ge(args.frequent_error_min)
    ].copy()

    all_protocol_wrong = subject_summary[
        subject_summary["n_protocols_wrong"].eq(
            subject_summary["n_protocols_available"]
        )
        & subject_summary["n_protocols_available"].gt(0)
    ].copy()

    seed_events = seed_error_events(seed_csv, task=args.task)

    # Save CSVs.
    events.to_csv(
        out / "01_all_misclassification_events.csv",
        index=False,
        encoding="utf-8-sig",
    )
    subject_summary.to_csv(
        out / "02_subject_cross_protocol_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )
    frequent.to_csv(
        out / "03_frequent_misclassified_subjects.csv",
        index=False,
        encoding="utf-8-sig",
    )
    all_protocol_wrong.to_csv(
        out / "04_wrong_in_all_available_protocols.csv",
        index=False,
        encoding="utf-8-sig",
    )
    transitions.to_csv(
        out / "05_error_transitions_by_protocol.csv",
        index=False,
        encoding="utf-8-sig",
    )
    psummary.to_csv(
        out / "06_protocol_error_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )
    class_summary.to_csv(
        out / "07_class_error_summary_by_protocol.csv",
        index=False,
        encoding="utf-8-sig",
    )
    center_summary.to_csv(
        out / "08_center_error_summary_by_protocol.csv",
        index=False,
        encoding="utf-8-sig",
    )
    seed_events.to_csv(
        out / "09_seed_level_error_events.csv",
        index=False,
        encoding="utf-8-sig",
    )

    readme = pd.DataFrame([
        [
            "Protocols",
            "gkf3, gkf5, gkf7, gkf10, loso, loco",
        ],
        [
            "Primary unit in 01_all_misclassification_events",
            "One row = one subject misclassified under one protocol",
        ],
        [
            "Primary unit in 02_subject_cross_protocol_summary",
            "One row = one subject, with six protocol-specific prediction/result columns",
        ],
        [
            "n_protocols_wrong",
            "How many of the available protocols misclassified this subject",
        ],
        [
            "wrong_protocols",
            "Exactly which experimental settings misclassified the subject",
        ],
        [
            "transition columns",
            "e.g. gkf5_transition = Moderate -> Mild",
        ],
        [
            "LOCO caveat",
            "LOCO may contain fewer evaluated subjects than GKF/LOSO because train-only centers are not tested.",
        ],
        [
            "Frequent error definition",
            f"n_protocols_wrong >= {args.frequent_error_min}",
        ],
    ], columns=["Item", "Definition"])

    xlsx = out / "task4_misclassification_all_protocols.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as w:
        events.to_excel(
            w, sheet_name="All_Error_Events", index=False
        )
        subject_summary.to_excel(
            w, sheet_name="Subject_Cross_Protocol", index=False
        )
        frequent.to_excel(
            w, sheet_name="Frequent_Errors", index=False
        )
        all_protocol_wrong.to_excel(
            w, sheet_name="Wrong_All_Available", index=False
        )
        transitions.to_excel(
            w, sheet_name="Transitions_By_Protocol", index=False
        )
        psummary.to_excel(
            w, sheet_name="Protocol_Summary", index=False
        )
        class_summary.to_excel(
            w, sheet_name="Class_By_Protocol", index=False
        )
        center_summary.to_excel(
            w, sheet_name="Center_By_Protocol", index=False
        )
        seed_events.to_excel(
            w, sheet_name="Seed_Error_Events", index=False
        )
        readme.to_excel(
            w, sheet_name="README", index=False
        )

    style_xlsx(xlsx)

    print("\n===== PROTOCOL ERROR SUMMARY =====")
    print(
        psummary[
            ["protocol", "n_subjects", "n_wrong", "accuracy", "error_rate"]
        ].to_string(index=False)
    )

    print("\n===== TOP ERROR TRANSITIONS BY PROTOCOL =====")
    if transitions.empty:
        print("No errors.")
    else:
        print(
            transitions[
                ["protocol", "error_transition", "n_errors"]
            ].to_string(index=False)
        )

    print("\n===== FREQUENTLY MISCLASSIFIED SUBJECTS =====")
    show = [
        c for c in [
            "subject_id", "name", "center", "cobb_angle",
            "true_label", "n_protocols_available",
            "n_protocols_wrong", "wrong_protocols",
            "most_common_error_transition",
            "stable_same_wrong_class",
        ]
        if c in frequent.columns
    ]
    print(
        frequent[show].to_string(index=False)
        if len(frequent)
        else "None."
    )

    print(f"\n[DONE] {xlsx}")


if __name__ == "__main__":
    main()
